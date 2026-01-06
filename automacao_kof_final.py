import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# 1. CONFIGURAÇÃO DE CAMINHOS
diretorio_atual = os.path.dirname(os.path.abspath(__file__))
# Substitua pelo nome exato do seu arquivo original se for diferente
caminho_entrada = os.path.join(diretorio_atual, "Kof 2k2 um 2019 Feb japanese ratio tier list.xlsx")
caminho_saida = os.path.join(diretorio_atual, "Kof_2k2_Rankeado_Final.xlsx")

# 2. PROCESSAMENTO DE DADOS (PANDAS)
df_kof = pd.read_excel(caminho_entrada, header=1)

# Lógica de pontuação e ordenação
df_kof['total_score'] = df_kof['point'] + df_kof['mid'] + df_kof['anchor']
df_sorted = df_kof.sort_values(by=['anchor', 'total_score'], ascending=[False, False]).copy()

# Classificação por Rank
classificacoes = []
for index, row in df_sorted.iterrows():
    val = row['anchor']
    if val > 7: rank = "Rank S"
    elif val == 7: rank = "Rank A"
    elif val == 6: rank = "Rank B"
    elif val == 5: rank = "Rank C"
    else: rank = "Rank D"
    classificacoes.append(rank)

df_sorted['Classificacao'] = classificacoes
df_final = df_sorted.drop(columns=['total_score'])
df_final.to_excel(caminho_saida, index=False)

# 3. ESTILIZAÇÃO E GRÁFICO (OPENPYXL)
wb = load_workbook(caminho_saida)
ws = wb.active

cores_rank = {
    'Rank S': 'FFD700', 'Rank A': 'C0C0C0', 'Rank B': 'CD7F32', 
    'Rank C': '87CEEB', 'Rank D': 'FF9999'
}
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Busca a coluna de Classificação
col_idx = None
for cell in ws[1]:
    if cell.value == "Classificacao":
        col_idx = cell.column
        break

# Aplicando estilos e cores
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    valor_rank = ws.cell(row=row[0].row, column=col_idx).value
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if cell.row == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", fill_type="solid")
        elif cell.column == col_idx:
            if valor_rank in cores_rank:
                cell.fill = PatternFill(start_color=cores_rank[valor_rank], fill_type="solid")
                cell.font = Font(bold=True)

# AJUSTE AUTOMÁTICO DE COLUNAS
for col in ws.columns:
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except: pass
    ws.column_dimensions[column_letter].width = max_length + 4

# 4. GERAÇÃO DO GRÁFICO
ordem_ranks = ["Rank S", "Rank A", "Rank B", "Rank C", "Rank D"]
todos_ranks = [ws.cell(row=i, column=col_idx).value for i in range(2, ws.max_row + 1)]

# Dados de apoio (I e J serão cobertos ou ficarão distantes)
ws["I1"], ws["J1"] = "Rank", "Qtd"
for i, rank in enumerate(ordem_ranks, start=2):
    ws.cell(row=i, column=9, value=rank)
    ws.cell(row=i, column=10, value=todos_ranks.count(rank))

chart = BarChart()
chart.title = "Distribuição de Ranks"
chart.legend = None
data = Reference(ws, min_col=10, min_row=1, max_row=6)
cats = Reference(ws, min_col=9, min_row=2, max_row=6)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# Colorir as barras
series = chart.series[0]
for i in range(len(ordem_ranks)):
    pt = DataPoint(idx=i)
    nome_rank = ws.cell(row=i+2, column=9).value
    pt.graphicalProperties.solidFill = cores_rank.get(nome_rank, "000000")
    series.dPt.append(pt)

# Posicionando o gráfico em G1 para ficar esteticamente agradável e ocultar os dados
ws.add_chart(chart, "G1") 

wb.save(caminho_saida)
print(f"✅ Automação concluída com sucesso! Arquivo: {caminho_saida}")