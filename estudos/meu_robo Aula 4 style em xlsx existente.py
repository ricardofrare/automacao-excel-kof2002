import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- SEU CÓDIGO EXISTENTE ---
diretorio_atual = os.path.dirname(os.path.abspath(__file__))
caminho_saida = os.path.join(diretorio_atual, "Kof_2k2_Rankeado.xlsx")

# (Aqui acontece o df_kof.to_excel(caminho_saida, index=False))

# --- INÍCIO DA MENTORIA DE ESTILIZAÇÃO ---

# 3. ESTILIZAÇÃO COM OPENPYXL
wb = load_workbook(caminho_saida)
ws = wb.active

# Configuração de Estilos
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Cores corrigidas para baterem com o seu texto exato
cores_rank = {
    'Rank S': 'FFD700', # Dourado
    'Rank A': 'C0C0C0', # Prata
    'Rank B': 'CD7F32', # Bronze
    'Rank C': 'E0E0E0', # Cinza Claro
    'Rank D': 'FF9999'  # Vermelho
}

# 4. Busca Dinâmica da Coluna "Classificacao"
# Isso evita erros se a coluna mudar de lugar
col_idx = None
for cell in ws[1]: # Varre o cabeçalho
    if cell.value == "Classificacao":
        col_idx = cell.column - 1
        break

# 5. Aplicação dos Estilos
if col_idx is not None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        # Valor da célula de Rank nesta linha
        valor_rank = row[col_idx].value
        
        for cell in row:
            # Borda e Alinhamento em todas as células (Padrão Profissional)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Formatação do Cabeçalho
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="333333", fill_type="solid")
            
            # Colorir APENAS a coluna de Rank (se não for o cabeçalho)
            elif cell.column == (col_idx + 1):
                if valor_rank in cores_rank:
                    cell.fill = PatternFill(start_color=cores_rank[valor_rank], fill_type="solid")
                    cell.font = Font(bold=True) # Destaca o texto do Rank

# 6. Auto-ajuste de largura das colunas
for col in ws.columns:
    max_length = 0
    column_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 4

# 7. Salvamento Final
wb.save(caminho_saida)

print(f"✅ Relatório pronto! Apenas a coluna de Rank foi colorida em: {caminho_saida}")