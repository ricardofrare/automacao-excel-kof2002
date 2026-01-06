import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# --- CONFIGURAÇÃO ---
diretorio_atual = os.path.dirname(os.path.abspath(__file__))
caminho_saida = os.path.join(diretorio_atual, "Kof_2k2_Rankeado.xlsx")

wb = load_workbook(caminho_saida)
ws = wb.active

# Cores e Ordem (Mantendo a lógica anterior)
cores_ranks = {
    "Rank S": "FFD700", "Rank A": "C0C0C0", 
    "Rank B": "CD7F32", "Rank C": "87CEEB", "Rank D": "FF9999"
}
ordem_desejada = ["Rank S", "Rank A", "Rank B", "Rank C", "Rank D"]

# 1. CAPTURA DE DADOS
todos_os_ranks = [ws.cell(row=i, column=5).value for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=5).value]

# 2. ESCREVENDO OS DADOS DE APOIO (Vamos deixar na I e J)
ws["I1"], ws["J1"] = "Rank", "Qtd"
linha_aux = 2
for rank in ordem_desejada:
    qtd = todos_os_ranks.count(rank)
    if qtd > 0:
        ws.cell(row=linha_aux, column=9, value=rank)
        ws.cell(row=linha_aux, column=10, value=qtd)
        linha_aux += 1

# 3. CONFIGURANDO O GRÁFICO
chart = BarChart()
chart.title = "Distribuição de Ranks"
chart.legend = None

# Referenciando as colunas I(9) e J(10)
data = Reference(ws, min_col=10, min_row=1, max_row=linha_aux - 1)
cats = Reference(ws, min_col=9, min_row=2, max_row=linha_aux - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# 4. COLORIR BARRAS
series = chart.series[0]
for i in range(linha_aux - 2):
    pt = DataPoint(idx=i)
    nome_rank = ws.cell(row=i+2, column=9).value
    pt.graphicalProperties.solidFill = cores_ranks.get(nome_rank, "000000")
    series.dPt.append(pt)

# 5. POSICIONAMENTO (Colocando o gráfico em cima dos dados!)
# Ao colocar o gráfico começando na célula I1, ele vai cobrir os números
ws.add_chart(chart, "I1") 

wb.save(caminho_saida)
print("Gráfico posicionado sobre os dados de apoio!")